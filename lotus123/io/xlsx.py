"""XLSX file format handler using openpyxl.

This module provides import and export functionality for Excel XLSX files,
with bidirectional translation to ensure round-trip fidelity.
"""

from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any, cast

from .xlsx_format_translator import FormatTranslator
from .xlsx_formula_translator import FormulaTranslator

if TYPE_CHECKING:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.styles import Alignment
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.worksheet import Worksheet

    from ..core.spreadsheet import Spreadsheet

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.worksheet import Worksheet

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Alignment prefix mapping
ALIGNMENT_TO_PREFIX = {
    "left": "'",
    "right": '"',
    "center": "^",
}

PREFIX_TO_ALIGNMENT = {
    "'": "left",
    '"': "right",
    "^": "center",
}


@dataclass
class XlsxImportWarnings:
    """Warnings collected during XLSX import."""

    sheet_count: int = 0
    imported_sheet_name: str = ""
    unsupported_formulas: list[tuple[str, str]] = field(default_factory=list)  # (cell_ref, formula)
    merged_cells: list[str] = field(default_factory=list)
    conditional_formatting: bool = False
    data_validations: bool = False

    def has_warnings(self) -> bool:
        """Check if there are any warnings."""
        return (
            self.sheet_count > 1
            or bool(self.unsupported_formulas)
            or bool(self.merged_cells)
            or self.conditional_formatting
            or self.data_validations
        )

    def to_message(self) -> str:
        """Generate human-readable warning message."""
        parts = []

        if self.sheet_count > 1:
            parts.append(
                f"Workbook has {self.sheet_count} sheets. "
                f"Imported '{self.imported_sheet_name}' only."
            )

        if self.unsupported_formulas:
            parts.append(f"{len(self.unsupported_formulas)} formulas have unsupported functions.")

        if self.merged_cells:
            parts.append(f"{len(self.merged_cells)} merged cell ranges were unmerged.")

        if self.conditional_formatting:
            parts.append("Conditional formatting was ignored.")

        if self.data_validations:
            parts.append("Data validations were ignored.")

        return " ".join(parts)


class XlsxReader:
    """Read XLSX files into spreadsheet."""

    def __init__(self, spreadsheet: Spreadsheet) -> None:
        """Initialize reader with target spreadsheet.

        Args:
            spreadsheet: Spreadsheet to load data into
        """
        self.spreadsheet = spreadsheet
        self.warnings = XlsxImportWarnings()

    @staticmethod
    def get_sheet_names(filepath: str) -> list[str]:
        """Get list of sheet names from an XLSX file.

        Args:
            filepath: Path to XLSX file

        Returns:
            List of sheet names

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX support. Install with: uv add openpyxl"
            )

        wb = load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    def load(self, filepath: str, sheet_name: str | None = None) -> XlsxImportWarnings:
        """Load XLSX file into spreadsheet.

        Args:
            filepath: Path to XLSX file
            sheet_name: Specific sheet to load (None = active sheet)

        Returns:
            Warnings object containing any issues encountered

        Raises:
            ImportError: If openpyxl is not installed
            ValueError: If file format is invalid
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX support. Install with: uv add openpyxl"
            )

        self.warnings = XlsxImportWarnings()
        self.spreadsheet.clear()

        wb = load_workbook(filepath, data_only=False)

        # Track sheet count
        self.warnings.sheet_count = len(wb.sheetnames)

        # Select sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif wb.active is not None:
            ws = wb.active
        else:
            # Fallback to first sheet if no active sheet
            ws = wb[wb.sheetnames[0]]

        self.warnings.imported_sheet_name = ws.title

        # Import cells
        self._import_cells(ws)

        # Import column widths
        self._import_column_widths(ws)

        # Import row heights
        self._import_row_heights(ws)

        # Import named ranges
        self._import_named_ranges(wb)

        # Import frozen panes
        self._import_frozen_panes(ws)

        # Check for unsupported features
        self._check_merged_cells(ws)
        self._check_conditional_formatting(ws)
        self._check_data_validations(ws)

        wb.close()

        self.spreadsheet.filename = ""  # Not native format
        self.spreadsheet.modified = True
        self.spreadsheet.rebuild_dependency_graph()

        return self.warnings

    def _import_cells(self, ws: Worksheet) -> None:
        """Import all cells from worksheet."""
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value is None:
                    continue

                row = cell.row - 1  # Convert to 0-based
                col = cell.column - 1

                # Determine if this is a formula
                # Note: Only treat as formula if data_type is "f" (formula)
                # Don't treat strings starting with "=" as formulas if data_type is "s"
                raw_value = ""

                if cell.data_type == "f":
                    # Formula cell
                    formula = cell.value if isinstance(cell.value, str) else f"={cell.value}"
                    lotus_formula = FormulaTranslator.excel_to_lotus(formula)

                    # Check for unsupported functions
                    unsupported = FormulaTranslator.get_unsupported_excel_functions(formula)
                    if unsupported:
                        cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                        self.warnings.unsupported_formulas.append((cell_ref, formula))

                    raw_value = lotus_formula
                else:
                    # Regular value - handle alignment for text
                    if isinstance(cell.value, str):
                        value = cell.value
                        # Check Excel alignment and add Lotus prefix
                        prefix = self._get_alignment_prefix(cell)
                        # For text starting with formula-like characters, always add ' prefix
                        # to prevent Lotus from interpreting as formula
                        if value and value[0] in ("=", "+", "-", "@") and not prefix:
                            prefix = "'"
                        raw_value = prefix + value
                    elif isinstance(cell.value, (int, float)):
                        raw_value = str(cell.value)
                    else:
                        raw_value = str(cell.value) if cell.value is not None else ""

                # Set cell value
                self.spreadsheet.set_cell(row, col, raw_value)

                # Import format code
                if cell.number_format and cell.number_format != "General":
                    lotus_format = FormatTranslator.excel_to_lotus(cell.number_format)
                    lotus_cell = self.spreadsheet.get_cell(row, col)
                    lotus_cell.format_code = lotus_format

    def _get_alignment_prefix(self, cell: Any) -> str:
        """Get Lotus alignment prefix from Excel cell alignment.

        Args:
            cell: openpyxl cell

        Returns:
            Lotus alignment prefix character or empty string
        """
        if cell.alignment and cell.alignment.horizontal:
            horizontal = cell.alignment.horizontal
            if horizontal in ALIGNMENT_TO_PREFIX:
                return ALIGNMENT_TO_PREFIX[horizontal]
        return ""

    def _import_column_widths(self, ws: Worksheet) -> None:
        """Import column widths."""
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width is not None and dim.width != 8.43:  # Default Excel width
                col_idx = column_index_from_string(col_letter) - 1
                # Excel width is in characters, roughly similar to Lotus
                self.spreadsheet.set_col_width(col_idx, int(dim.width))

    def _import_row_heights(self, ws: Worksheet) -> None:
        """Import row heights."""
        for row_num, dim in ws.row_dimensions.items():
            if dim.height is not None and dim.height != 15:  # Default Excel height
                row_idx = row_num - 1
                # Excel height is in points, Lotus uses lines (roughly 15pt = 1 line)
                lines = max(1, round(dim.height / 15))
                self.spreadsheet.set_row_height(row_idx, lines)

    def _import_named_ranges(self, wb: Workbook) -> None:
        """Import named ranges from workbook."""
        # Iterate over defined names - iteration yields name strings
        for name_key in wb.defined_names:
            defined_name = wb.defined_names[name_key]
            if defined_name.value and not defined_name.value.startswith("#"):
                try:
                    # Parse the reference
                    ref_str = defined_name.value
                    # Remove sheet name if present
                    if "!" in ref_str:
                        ref_str = ref_str.split("!")[-1]
                    ref_str = ref_str.replace("$", "")  # Remove absolute markers

                    self.spreadsheet.named_ranges.add_from_string(defined_name.name, ref_str)
                except (ValueError, KeyError, AttributeError):
                    pass  # Skip invalid names

    def _import_frozen_panes(self, ws: Worksheet) -> None:
        """Import frozen pane settings."""
        if ws.freeze_panes:
            # freeze_panes is like "B2" meaning row 1 and col A are frozen
            freeze_cell = ws.freeze_panes
            col_letter = "".join(c for c in freeze_cell if c.isalpha())
            row_num = "".join(c for c in freeze_cell if c.isdigit())

            if col_letter:
                self.spreadsheet.frozen_cols = column_index_from_string(col_letter) - 1
            if row_num:
                self.spreadsheet.frozen_rows = int(row_num) - 1

    def _check_merged_cells(self, ws: Worksheet) -> None:
        """Check for merged cells (not fully supported)."""
        for merge_range in ws.merged_cells.ranges:
            self.warnings.merged_cells.append(str(merge_range))

    def _check_conditional_formatting(self, ws: Worksheet) -> None:
        """Check for conditional formatting."""
        # Check if there are any conditional formatting rules
        if len(ws.conditional_formatting) > 0:
            self.warnings.conditional_formatting = True

    def _check_data_validations(self, ws: Worksheet) -> None:
        """Check for data validations."""
        if ws.data_validations.dataValidation:
            self.warnings.data_validations = True


class XlsxWriter:
    """Write spreadsheet to XLSX format."""

    def __init__(self, spreadsheet: Spreadsheet) -> None:
        """Initialize writer with source spreadsheet.

        Args:
            spreadsheet: Spreadsheet to export
        """
        self.spreadsheet = spreadsheet

    def save(self, filepath: str) -> None:
        """Save spreadsheet to XLSX file.

        Args:
            filepath: Path to save file

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX support. Install with: uv add openpyxl"
            )

        wb = Workbook()
        ws = wb.active
        assert ws is not None, "New workbook should have an active sheet"
        ws.title = "Sheet1"

        # Export cells
        self._export_cells(ws)

        # Export column widths
        self._export_column_widths(ws)

        # Export row heights
        self._export_row_heights(ws)

        # Export named ranges
        self._export_named_ranges(wb)

        # Export frozen panes
        self._export_frozen_panes(ws)

        wb.save(filepath)
        wb.close()

    def _export_cells(self, ws: Worksheet) -> None:
        """Export all cells to worksheet."""
        from ..core.cell import ALIGNMENT_PREFIXES

        for (row, col), cell in self.spreadsheet.cells.items():
            if cell.is_empty:
                continue

            excel_row = row + 1  # Convert to 1-based
            excel_col = col + 1
            # Cast to Cell since we're creating new cells, not accessing merged cells
            excel_cell = cast("Cell", ws.cell(row=excel_row, column=excel_col))

            if cell.is_formula:
                # Convert formula
                excel_formula = FormulaTranslator.lotus_to_excel(cell.raw_value)

                # Check if this looks like a valid Excel formula
                # Invalid patterns that should be exported as text:
                # - === (text separator)
                # - =@ (malformed, @ should not follow =)
                # - == (double equals)
                # - =!= or similar invalid syntax
                is_valid_formula = True
                if excel_formula.startswith("==="):
                    is_valid_formula = False
                elif excel_formula.startswith("=@"):
                    is_valid_formula = False
                elif excel_formula.startswith("=="):
                    is_valid_formula = False
                elif len(excel_formula) >= 2 and excel_formula[1] in "!<>":
                    # Things like =!= are invalid
                    is_valid_formula = False

                if is_valid_formula:
                    excel_cell.value = excel_formula
                else:
                    # Export as text with ' prefix to show it's not a formula
                    excel_cell.value = excel_formula
                    excel_cell.data_type = "s"
            else:
                # Regular value - handle alignment prefix
                raw = cell.raw_value
                alignment_prefix = None

                if raw and raw[0] in ALIGNMENT_PREFIXES:
                    alignment_prefix = raw[0]
                    display_value = raw[1:]
                else:
                    display_value = raw

                # If there's an explicit alignment prefix, treat as text
                # (user intentionally marked this as a label, e.g., '2023 means text "2023")
                if alignment_prefix:
                    excel_cell.value = display_value
                    if display_value and display_value[0] in ("=", "+", "-", "@"):
                        excel_cell.data_type = "s"  # Explicitly set as string
                else:
                    # Try to convert to number (no prefix = auto-detect type)
                    try:
                        # Check for integer
                        if "." not in display_value and "e" not in display_value.lower():
                            excel_cell.value = int(display_value)
                        else:
                            excel_cell.value = float(display_value)
                    except (ValueError, TypeError):
                        # For text that starts with formula-like characters,
                        # set value and mark as string to prevent interpretation as formula
                        excel_cell.value = display_value
                        if display_value and display_value[0] in ("=", "+", "-", "@"):
                            excel_cell.data_type = "s"  # Explicitly set as string

                # Apply alignment
                if alignment_prefix and alignment_prefix in PREFIX_TO_ALIGNMENT:
                    excel_cell.alignment = Alignment(
                        horizontal=PREFIX_TO_ALIGNMENT[alignment_prefix]
                    )

            # Apply format
            if cell.format_code and cell.format_code != "G":
                excel_format = FormatTranslator.lotus_to_excel(cell.format_code)
                excel_cell.number_format = excel_format

    def _export_column_widths(self, ws: Worksheet) -> None:
        """Export column widths."""
        for col, width in self.spreadsheet.col_widths.items():
            col_letter = get_column_letter(col + 1)
            ws.column_dimensions[col_letter].width = width

    def _export_row_heights(self, ws: Worksheet) -> None:
        """Export row heights."""
        for row, height in self.spreadsheet.row_heights.items():
            # Convert Lotus lines to Excel points (1 line = ~15 points)
            ws.row_dimensions[row + 1].height = height * 15

    def _export_named_ranges(self, wb: Workbook) -> None:
        """Export named ranges."""
        from ..core.reference import CellReference, RangeReference

        for named in self.spreadsheet.named_ranges.list_all():
            ref = named.reference

            # Build Excel-style reference
            if isinstance(ref, CellReference):
                excel_ref = f"Sheet1!${get_column_letter(ref.col + 1)}${ref.row + 1}"
            elif isinstance(ref, RangeReference):
                start = f"${get_column_letter(ref.start.col + 1)}${ref.start.row + 1}"
                end = f"${get_column_letter(ref.end.col + 1)}${ref.end.row + 1}"
                excel_ref = f"Sheet1!{start}:{end}"
            else:
                continue

            defined_name = DefinedName(named.name, attr_text=excel_ref)
            wb.defined_names.add(defined_name)

    def _export_frozen_panes(self, ws: Worksheet) -> None:
        """Export frozen pane settings."""
        if self.spreadsheet.frozen_rows > 0 or self.spreadsheet.frozen_cols > 0:
            freeze_col = get_column_letter(self.spreadsheet.frozen_cols + 1)
            freeze_row = self.spreadsheet.frozen_rows + 1
            ws.freeze_panes = f"{freeze_col}{freeze_row}"


# Convenience functions
def load_xlsx(
    spreadsheet: Spreadsheet, filepath: str, sheet_name: str | None = None
) -> XlsxImportWarnings:
    """Load XLSX file into spreadsheet.

    Args:
        spreadsheet: Target spreadsheet
        filepath: Path to XLSX file
        sheet_name: Specific sheet to load (None = active sheet)

    Returns:
        Warnings object
    """
    reader = XlsxReader(spreadsheet)
    return reader.load(filepath, sheet_name)


def save_xlsx(spreadsheet: Spreadsheet, filepath: str) -> None:
    """Save spreadsheet to XLSX file.

    Args:
        spreadsheet: Source spreadsheet
        filepath: Path to save file
    """
    writer = XlsxWriter(spreadsheet)
    writer.save(filepath)


def get_xlsx_sheet_names(filepath: str) -> list[str]:
    """Get list of sheet names from XLSX file.

    Args:
        filepath: Path to XLSX file

    Returns:
        List of sheet names
    """
    return XlsxReader.get_sheet_names(filepath)
